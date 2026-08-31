import os
import sys
import asyncio
import logging
import openpyxl
from datetime import datetime, timedelta
from dotenv import load_dotenv
from playwright.async_api import async_playwright

# Carrega configurações do arquivo .env
load_dotenv()

# Configuração de Logs
log_file = os.getenv("LOG_FILE", "pwr_sales_check.log")
log_level_str = os.getenv("LOG_LEVEL", "INFO").upper()
log_level = getattr(logging, log_level_str, logging.INFO)

logging.basicConfig(
    level=log_level,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)

# Cria pastas necessárias
os.makedirs("screenshots", exist_ok=True)
os.makedirs("reports", exist_ok=True)

# Credenciais e Configurações
USER = os.getenv("DOMINOS_PWR_USER")
PASSWORD = os.getenv("DOMINOS_PWR_PASSWORD")
HEADLESS_STR = os.getenv("HEADLESS", "False").lower()
HEADLESS = HEADLESS_STR == "true"

def load_store_mappings():
    """
    Carrega o mapeamento de ID de loja para Nome da Loja, Consultor e Tipo (Franquia/Própria)
    a partir das planilhas Excel presentes na pasta 'lojas com id nome e consultor'.
    """
    folder = "lojas com id nome e consultor"
    mapping = {}
    
    # 1. Ler lojas próprias
    path_proprias = os.path.join(folder, "lojas proprias dominos.xlsx")
    if os.path.exists(path_proprias):
        try:
            wb = openpyxl.load_workbook(path_proprias, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    store_id = str(row[0]).strip()
                    store_name = str(row[1]).strip() if row[1] else "N/A"
                    consultant = str(row[2]).strip() if row[2] else "N/A"
                    mapping[store_id] = {
                        "name": store_name,
                        "consultant": consultant,
                        "type": "Própria"
                    }
            logging.info(f"Mapeamento: {len(mapping)} lojas próprias carregadas.")
        except Exception as e:
            logging.error(f"Erro ao ler lojas próprias: {e}")
    else:
        logging.warning(f"Planilha de lojas próprias não encontrada em {path_proprias}")
            
    # 2. Ler lojas franquias
    path_franquias = os.path.join(folder, "lojas_por_consultor_mapeado_v2.xlsx")
    if os.path.exists(path_franquias):
        try:
            wb = openpyxl.load_workbook(path_franquias, data_only=True)
            # A planilha de interesse é 'lojas por consultor'
            sheet = wb['lojas por consultor'] if 'lojas por consultor' in wb.sheetnames else wb.active
            count_franquias = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    store_id = str(row[0]).strip()
                    store_name = str(row[1]).strip() if row[1] else "N/A"
                    consultant = str(row[3]).strip() if row[3] else "N/A"
                    # Sobrescreve ou adiciona
                    mapping[store_id] = {
                        "name": store_name,
                        "consultant": consultant,
                        "type": "Franquia"
                    }
                    count_franquias += 1
            logging.info(f"Mapeamento: {count_franquias} lojas de franquias carregadas. Total mapeado: {len(mapping)} lojas.")
        except Exception as e:
            logging.error(f"Erro ao ler lojas franquias: {e}")
    else:
        logging.warning(f"Planilha de lojas franquias não encontrada em {path_franquias}")
            
    return mapping

def get_report_dates():
    """
    Calcula as datas necessárias para o relatório.
    D-1 é ontem. O período do mês vai do dia 01 do mês de ontem até ontem.
    """
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    
    begin_date_str = yesterday.strftime("01/%m/%Y")
    end_date_str = yesterday.strftime("%d/%m/%Y")
    
    return yesterday.strftime("%Y_%m_%d"), begin_date_str, end_date_str

async def parse_main_table(page):
    """
    Extrai as informações de lojas e o valor da Coluna E (dias sem subida de vendas).
    """
    table_id = "#ASPxGridViewMainReport_DXMainTable"
    try:
        await page.wait_for_selector(table_id, timeout=15000)
    except Exception as e:
        logging.error(f"Tabela principal {table_id} não carregou ou não foi encontrada: {e}")
        return None
        
    table_locator = page.locator(table_id)
    rows = await table_locator.locator("tbody tr, tr").all()
    
    data = {}
    for row in rows:
        # Ignora cabeçalhos
        if await row.locator("th").count() > 0:
            continue
            
        cells = await row.locator("td").all()
        row_data = []
        for cell in cells:
            row_data.append((await cell.inner_text()).strip())
            
        if row_data and row_data[0].isdigit():
            store_id = row_data[0]
            col_e_val = row_data[1]
            
            # Se col_e_val for um número (como '1', '10'), convertemos para int, senão é 0
            days_missing = int(col_e_val) if col_e_val.isdigit() else 0
            data[store_id] = days_missing
            
    return data

def generate_html_dashboard(yesterday_missing, monthly_missing, store_mapping, begin_date_str, end_date_str, total_stores):
    """
    Gera um painel interativo HTML com as informações de não subida de vendas,
    cruzando dados de nome de loja e consultores, permitindo busca e filtros.
    """
    # Preparar dados para ontem (D-1)
    d1_list = []
    for store_id in yesterday_missing:
        meta = store_mapping.get(store_id, {"name": "N/A", "consultant": "N/A", "type": "N/A"})
        d1_list.append({
            "id": store_id,
            "name": meta["name"],
            "consultant": meta["consultant"],
            "type": meta["type"]
        })
        
    # Preparar dados para o mês
    month_list = []
    for store_id, days in monthly_missing.items():
        meta = store_mapping.get(store_id, {"name": "N/A", "consultant": "N/A", "type": "N/A"})
        month_list.append({
            "id": store_id,
            "name": meta["name"],
            "consultant": meta["consultant"],
            "type": meta["type"],
            "days": days
        })
    # Ordenar por dias em atraso decrescente
    month_list.sort(key=lambda x: x["days"], reverse=True)

    # Obter lista de consultores únicos para o filtro
    all_consultants = sorted(list(set(
        [m["consultant"] for m in store_mapping.values()] + 
        [d["consultant"] for d in d1_list] + 
        [m["consultant"] for m in month_list]
    )))
    all_consultants = [c for c in all_consultants if c and c != "N/A"]
    import base64
    
    # 1. Carrega o logo horizontal para o header
    logo_svg = ""
    path_logo = r"logos_svg/DPZ_2025_Logo_CombinationMark_Horizontal_Blue_RGB.svg"
    if os.path.exists(path_logo):
        try:
            with open(path_logo, "r", encoding="utf-8") as f:
                logo_svg = f.read()
            # Remove xml declaration
            if logo_svg.startswith("<?xml"):
                logo_svg = logo_svg[logo_svg.find("?>")+2:].strip()
            # Injeta classe para dimensionamento no Tailwind
            logo_svg = logo_svg.replace("<svg", '<svg class="h-10 w-auto"')
        except Exception as e:
            logging.error(f"Erro ao carregar o logo SVG para o painel: {e}")
            
    if not logo_svg:
        logo_svg = '<h1 class="text-2xl font-bold text-slate-900 tracking-tight">PWR Domino\'s</h1>'

    # 2. Carrega a tile para o favicon
    favicon_link = ""
    path_tile = r"logos_svg/DPZ_2025_Logo_Tile_Blue_RGB.svg"
    if os.path.exists(path_tile):
        try:
            with open(path_tile, "rb") as f:
                b64_tile = base64.b64encode(f.read()).decode("utf-8")
            favicon_link = f'<link rel="icon" type="image/svg+xml" href="data:image/svg+xml;base64,{b64_tile}">'
        except Exception as e:
            logging.error(f"Erro ao carregar o favicon SVG: {e}")

    import json
    d1_json = json.dumps(d1_list)
    month_json = json.dumps(month_list)

    html_template = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PWR Domino's - Painel de Subida de Vendas</title>
    {favicon_link}
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        body {{
            font-family: 'Inter', sans-serif;
        }}
    </style>
</head>
<body class="bg-gray-50 min-h-screen text-gray-800">
    <div class="max-w-7xl mx-auto px-4 py-8">
        
        <!-- Header -->
        <header class="flex flex-col md:flex-row md:items-center md:justify-between pb-6 mb-8 border-b border-gray-200">
            <div>
                <div class="flex items-center space-x-3 mb-2">
                    <div class="h-10 w-auto flex items-center">
                        {logo_svg}
                    </div>
                </div>
                <p class="text-xs text-gray-500">Painel diário de controle de subida de vendas</p>
            </div>
            <div class="mt-4 md:mt-0 text-left md:text-right bg-white p-3 rounded-lg shadow-sm border border-gray-100">
                <p class="text-xs text-gray-400">Última atualização</p>
                <p class="text-sm font-semibold text-slate-800">{datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
                <p class="text-xs text-gray-500 mt-1">Período analisado: <span class="font-medium text-slate-700">{begin_date_str} a {end_date_str}</span></p>
            </div>
        </header>

        <!-- Cards de Estatísticas -->
        <div class="grid grid-cols-1 md:grid-cols-3 gap-6 mb-8">
            <div class="bg-white rounded-xl shadow-sm border border-gray-100 p-6">
                <div class="flex items-center justify-between">
                    <div>
                        <p class="text-xs font-semibold text-gray-400 uppercase tracking-wider">Total de Lojas Analisadas</p>
                        <p class="text-3xl font-bold text-slate-950 mt-1">{total_stores}</p>
                    </div>
                    <div class="p-3 bg-blue-50 text-blue-600 rounded-lg">
                        <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M19 21V5a2 2 0 00-2-2H7a2 2 0 00-2 2v16m14 0h2m-2 0h-5m-9 0H3m2 0h5M9 7h1m-1 4h1m4-4h1m-1 4h1m-5 10v-5a1 1 0 011-1h2a1 1 0 011 1v5m-4 0h4" /></svg>
                    </div>
                </div>
            </div>

            <div class="bg-white rounded-xl shadow-sm border border-gray-100 p-6">
                <div class="flex items-center justify-between">
                    <div>
                        <p class="text-xs font-semibold text-gray-400 uppercase tracking-wider">Pendentes de Ontem (D-1)</p>
                        <p class="text-3xl font-bold {'text-rose-600' if d1_list else 'text-green-600'} mt-1">{len(d1_list)}</p>
                    </div>
                    <div class="p-3 {'bg-rose-50 text-rose-600' if d1_list else 'bg-green-50 text-green-600'} rounded-lg">
                        <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-3L13.732 4c-.77-1.333-2.694-1.333-3.464 0L3.34 16c-.77 1.333.192 3 1.732 3z" /></svg>
                    </div>
                </div>
                <p class="text-xs text-gray-500 mt-2">Refere-se ao dia {end_date_str}</p>
            </div>

            <div class="bg-white rounded-xl shadow-sm border border-gray-100 p-6">
                <div class="flex items-center justify-between">
                    <div>
                        <p class="text-xs font-semibold text-gray-400 uppercase tracking-wider">Pendentes no Mês Acumulado</p>
                        <p class="text-3xl font-bold {'text-amber-600' if month_list else 'text-green-600'} mt-1">{len(month_list)}</p>
                    </div>
                    <div class="p-3 {'bg-amber-50 text-amber-600' if month_list else 'bg-green-50 text-green-600'} rounded-lg">
                        <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M8 7V3m8 4V3m-9 8h10M5 21h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v12a2 2 0 002 2z" /></svg>
                    </div>
                </div>
                <p class="text-xs text-gray-500 mt-2">Total de lojas com algum dia faltante no mês</p>
            </div>
        </div>

        <!-- Filtros Interativos -->
        <div class="bg-white rounded-xl shadow-sm border border-gray-100 p-4 mb-8 flex flex-col md:flex-row gap-4 items-center justify-between">
            <div class="w-full md:w-1/3">
                <label class="block text-xs font-semibold text-gray-400 uppercase mb-1">Buscar Loja</label>
                <input type="text" id="searchInput" placeholder="Digite ID ou nome da loja..." class="w-full bg-gray-50 border border-gray-200 rounded-lg px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-blue-500">
            </div>
            
            <div class="w-full md:w-1/4">
                <label class="block text-xs font-semibold text-gray-400 uppercase mb-1">Filtrar por Consultor</label>
                <select id="consultantFilter" class="w-full bg-gray-50 border border-gray-200 rounded-lg px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-blue-500">
                    <option value="ALL">Todos os Consultores</option>
                    {"".join(f'<option value="{c}">{c}</option>' for c in all_consultants)}
                </select>
            </div>

            <div class="w-full md:w-1/4">
                <label class="block text-xs font-semibold text-gray-400 uppercase mb-1">Filtrar por Tipo</label>
                <select id="typeFilter" class="w-full bg-gray-50 border border-gray-200 rounded-lg px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-blue-500">
                    <option value="ALL">Todos os Tipos</option>
                    <option value="Franquia">Franquia</option>
                    <option value="Própria">Própria</option>
                </select>
            </div>
        </div>

        <!-- Abas (Tabs) -->
        <div class="mb-6 border-b border-gray-200 flex space-x-6">
            <button id="tabD1Btn" class="pb-3 text-sm font-semibold border-b-2 border-blue-600 text-blue-600 focus:outline-none" onclick="switchTab('D1')">
                Lojas Pendentes Ontem (D-1) <span class="ml-1.5 px-2 py-0.5 bg-rose-100 text-rose-800 text-xs rounded-full font-bold">{len(d1_list)}</span>
            </button>
            <button id="tabMonthBtn" class="pb-3 text-sm font-semibold border-b-2 border-transparent text-gray-500 hover:text-gray-700 focus:outline-none" onclick="switchTab('Month')">
                Lojas Pendentes no Mês <span class="ml-1.5 px-2 py-0.5 bg-amber-100 text-amber-800 text-xs rounded-full font-bold">{len(month_list)}</span>
            </button>
        </div>

        <!-- Tabela D-1 -->
        <div id="tableD1Container" class="bg-white rounded-xl shadow-sm border border-gray-100 overflow-hidden">
            <table class="min-w-full divide-y divide-gray-200">
                <thead class="bg-gray-50">
                    <tr>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">ID Loja</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Nome da Loja</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Consultor / Coordenador</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Tipo</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Atraso</th>
                    </tr>
                </thead>
                <tbody id="tableD1Body" class="bg-white divide-y divide-gray-200 text-sm">
                    <!-- Dinâmico -->
                </tbody>
            </table>
            <div id="noDataD1" class="hidden p-8 text-center text-gray-500">Nenhuma loja pendente ontem.</div>
        </div>

        <!-- Tabela Mês -->
        <div id="tableMonthContainer" class="bg-white rounded-xl shadow-sm border border-gray-100 overflow-hidden hidden">
            <table class="min-w-full divide-y divide-gray-200">
                <thead class="bg-gray-50">
                    <tr>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">ID Loja</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Nome da Loja</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Consultor / Coordenador</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider">Tipo</th>
                        <th class="px-6 py-3 text-left text-xs font-semibold text-gray-500 uppercase tracking-wider text-center">Dias sem Vendas</th>
                    </tr>
                </thead>
                <tbody id="tableMonthBody" class="bg-white divide-y divide-gray-200 text-sm">
                    <!-- Dinâmico -->
                </tbody>
            </table>
            <div id="noDataMonth" class="hidden p-8 text-center text-gray-500">Nenhuma loja com pendência acumulada no mês.</div>
        </div>

    </div>

    <!-- Script de Filtro -->
    <script>
        const d1Data = {d1_json};
        const monthData = {month_json};
        let currentTab = 'D1';

        function switchTab(tab) {{
            currentTab = tab;
            const tabD1Btn = document.getElementById('tabD1Btn');
            const tabMonthBtn = document.getElementById('tabMonthBtn');
            const tableD1 = document.getElementById('tableD1Container');
            const tableMonth = document.getElementById('tableMonthContainer');

            if (tab === 'D1') {{
                tabD1Btn.classList.add('border-blue-600', 'text-blue-600');
                tabD1Btn.classList.remove('border-transparent', 'text-gray-500');
                tabMonthBtn.classList.remove('border-blue-600', 'text-blue-600');
                tabMonthBtn.classList.add('border-transparent', 'text-gray-500');
                tableD1.classList.remove('hidden');
                tableMonth.classList.add('hidden');
            }} else {{
                tabMonthBtn.classList.add('border-blue-600', 'text-blue-600');
                tabMonthBtn.classList.remove('border-transparent', 'text-gray-500');
                tabD1Btn.classList.remove('border-blue-600', 'text-blue-600');
                tabD1Btn.classList.add('border-transparent', 'text-gray-500');
                tableMonth.classList.remove('hidden');
                tableD1.classList.add('hidden');
            }}
            filterData();
        }}

        function filterData() {{
            const query = document.getElementById('searchInput').value.toLowerCase().trim();
            const consultant = document.getElementById('consultantFilter').value;
            const type = document.getElementById('typeFilter').value;

            if (currentTab === 'D1') {{
                const tbody = document.getElementById('tableD1Body');
                tbody.innerHTML = '';
                
                const filtered = d1Data.filter(item => {{
                    const matchQuery = item.id.includes(query) || item.name.toLowerCase().includes(query);
                    const matchConsultant = consultant === 'ALL' || item.consultant === consultant;
                    const matchType = type === 'ALL' || item.type === type;
                    return matchQuery && matchConsultant && matchType;
                }});

                if (filtered.length === 0) {{
                    document.getElementById('noDataD1').classList.remove('hidden');
                }} else {{
                    document.getElementById('noDataD1').classList.add('hidden');
                    filtered.forEach(item => {{
                        const tr = document.createElement('tr');
                        tr.innerHTML = `
                            <td class="px-6 py-4 whitespace-nowrap font-medium text-gray-900">${{item.id}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-gray-600">${{item.name}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-gray-600">${{item.consultant}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-xs">
                                <span class="px-2.5 py-1 rounded-full font-semibold ${{item.type === 'Própria' ? 'bg-indigo-50 text-indigo-700' : 'bg-orange-50 text-orange-700'}}">${{item.type}}</span>
                            </td>
                            <td class="px-6 py-4 whitespace-nowrap text-xs font-bold text-rose-600">
                                <span class="bg-rose-50 px-2.5 py-1 rounded-lg">1 dia</span>
                            </td>
                        `;
                        tbody.appendChild(tr);
                    }});
                }}
            }} else {{
                const tbody = document.getElementById('tableMonthBody');
                tbody.innerHTML = '';
                
                const filtered = monthData.filter(item => {{
                    const matchQuery = item.id.includes(query) || item.name.toLowerCase().includes(query);
                    const matchConsultant = consultant === 'ALL' || item.consultant === consultant;
                    const matchType = type === 'ALL' || item.type === type;
                    return matchQuery && matchConsultant && matchType;
                }});

                if (filtered.length === 0) {{
                    document.getElementById('noDataMonth').classList.remove('hidden');
                }} else {{
                    document.getElementById('noDataMonth').classList.add('hidden');
                    filtered.forEach(item => {{
                        // Determinar a cor com base nos dias em atraso
                        const severityClass = item.days >= 5 ? 'text-red-700 bg-red-50' : (item.days >= 2 ? 'text-amber-700 bg-amber-50' : 'text-yellow-700 bg-yellow-50');
                        tr = document.createElement('tr');
                        tr.innerHTML = `
                            <td class="px-6 py-4 whitespace-nowrap font-medium text-gray-900">${{item.id}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-gray-600">${{item.name}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-gray-600">${{item.consultant}}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-xs">
                                <span class="px-2.5 py-1 rounded-full font-semibold ${{item.type === 'Própria' ? 'bg-indigo-50 text-indigo-700' : 'bg-orange-50 text-orange-700'}}">${{item.type}}</span>
                            </td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-xs font-bold">
                                <span class="px-3 py-1 rounded-lg ${{severityClass}}">${{item.days}} dia(s)</span>
                            </td>
                        `;
                        tbody.appendChild(tr);
                    }});
                }}
            }}
        }}

        // Escutas de eventos para filtros
        document.getElementById('searchInput').addEventListener('input', filterData);
        document.getElementById('consultantFilter').addEventListener('change', filterData);
        document.getElementById('typeFilter').addEventListener('change', filterData);

        // Inicializar dados
        filterData();
    </script>
</body>
</html>
"""
    with open("painel_vendas.html", "w", encoding="utf-8") as f:
        f.write(html_template)
    logging.info("Painel HTML de vendas gerado com sucesso em painel_vendas.html")

async def wait_for_loading_to_finish(page):
    """
    Aguarda até que os painéis de carregamento da DevExtreme / PWR fiquem ocultos (hidden).
    """
    logging.info("Aguardando finalização do carregamento (loading panel)...")
    await page.wait_for_timeout(1500) # Pequena espera estável para dar tempo do loader surgir
    
    loaders = [
        page.locator("text=Loading"),
        page.locator(".dxlpLoadingPanel"),
        page.locator(".dxpcLoadingPanel"),
        page.locator(".dx-loadpanel"),
        page.locator(".dx-loadindicator")
    ]
    
    for loader in loaders:
        try:
            if await loader.count() > 0:
                await loader.wait_for(state="hidden", timeout=55000)
        except Exception:
            pass
            
    await page.wait_for_timeout(1500) # Margem de segurança pós-carregamento

async def run_automation():
    if not USER or USER == "seu_usuario_aqui" or not PASSWORD or PASSWORD == "sua_senha_aqui":
        logging.error("Credenciais não configuradas no arquivo .env!")
        sys.exit(1)
        
    # Carregar planilhas de mapeamento
    store_mapping = load_store_mappings()
        
    date_suffix, begin_date_str, end_date_str = get_report_dates()
    logging.info(f"Iniciando verificação de vendas PWR. Período do mês: {begin_date_str} até {end_date_str}")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized"
            ]
        )
        
        context = await browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        page = await context.new_page()
        url = "https://pwr.dominos.com"
        
        try:
            # 1. Login no Portal
            logging.info("Acessando portal...")
            await page.goto(url, wait_until="networkidle", timeout=60000)
            
            logging.info("Preenchendo credenciais...")
            await page.locator("#txtUsername").fill(USER)
            await page.locator("#txtPassword").fill(PASSWORD)
            
            logging.info("Clicando em Sign In...")
            await page.locator("#dxLoginButton").click()
            
            logging.info("Aguardando carregamento da página principal...")
            await page.wait_for_timeout(10000) # Espera estável pós-login
            
            # Validação do Login
            current_url = page.url
            if "Login.aspx" in current_url:
                error_msg = ""
                if await page.locator("#errorLabel").is_visible():
                    error_msg = await page.locator("#errorLabel").inner_text()
                logging.error(f"Erro no login: '{error_msg}'")
                await page.screenshot(path="screenshots/erro_login.png")
                return False
                
            logging.info("Login bem-sucedido!")
            
            # --- MUDANÇA DE ESCOPO: Selecionar All Stores (Stores) ---
            logging.info("Alterando escopo de visualização para 'All Stores (Stores)'...")
            await page.locator(".pwr_scope_selection_text").first.click()
            await page.wait_for_timeout(2000)
            
            # Clica em All Stores (Stores) na árvore/menu
            all_stores_option = page.get_by_text("All Stores (Stores)", exact=True)
            await all_stores_option.click()
            
            logging.info("Aguardando atualização pós mudança de escopo (All Stores)...")
            await wait_for_loading_to_finish(page)
            
            # --- MOVIMENTO 1: Verificar ontem (D-1) ---
            logging.info("MOVIMENTO 1: Analisando vendas de ontem (D-1)...")
            await page.screenshot(path="screenshots/02_keys_summary_ontem.png")
            
            yesterday_data = await parse_main_table(page)
            if yesterday_data is None:
                logging.error("Não foi possível ler os dados de ontem.")
                return False
                
            total_stores_count = len(yesterday_data)
            logging.info(f"Total de lojas carregadas: {total_stores_count}")
                
            # Identificar lojas que não subiram ontem (dias em atraso >= 1)
            stores_missing_yesterday = [store for store, days in yesterday_data.items() if days >= 1]
            logging.info(f"Lojas sem vendas de ontem ({len(stores_missing_yesterday)}): {stores_missing_yesterday}")
            
            # --- MOVIMENTO 2: Verificar do dia 1 do mês até ontem (D-1) ---
            logging.info(f"MOVIMENTO 2: Configurando intervalo customizado ({begin_date_str} a {end_date_str})...")
            
            # Abre o menu de datas
            await page.locator(".dx-icon-clock").first.click()
            await page.wait_for_timeout(2000)
            
            # Clica em Custom
            custom_option = page.locator(".dx-menu-item-text").filter(has_text="Custom")
            await custom_option.first.click()
            await page.wait_for_timeout(3000)
            
            # Preenche as datas nos campos
            begin_input = page.locator("#custom_date_begin_selector input.dx-texteditor-input")
            end_input = page.locator("#custom_date_end_selector input.dx-texteditor-input")
            
            await begin_input.focus()
            await begin_input.fill(begin_date_str)
            await page.wait_for_timeout(1000)
            
            await end_input.focus()
            await end_input.fill(end_date_str)
            await page.wait_for_timeout(1000)
            
            # Clica fora para atualizar a grid
            await page.locator(".pwrJSMenuHeaderLong").first.click()
            logging.info("Aguardando recarregamento da tabela de vendas...")
            await wait_for_loading_to_finish(page)
            
            await page.screenshot(path="screenshots/03_keys_summary_mes.png")
            
            monthly_data = await parse_main_table(page)
            if monthly_data is None:
                logging.error("Não foi possível ler os dados acumulados do mês.")
                return False
                
            # Identificar lojas com dias pendentes no mês (dias sem subida >= 1)
            stores_missing_month = {store: days for store, days in monthly_data.items() if days >= 1}
            logging.info(f"Lojas com dias pendentes no mês ({len(stores_missing_month)}): {stores_missing_month}")
            
            # --- GERAR RELATÓRIO TXT ---
            report_lines = []
            report_lines.append("="*80)
            report_lines.append(f"RELATÓRIO DE VERIFICAÇÃO DE VENDAS PWR - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            report_lines.append("="*80)
            report_lines.append(f"Período Mensal Analisado: {begin_date_str} a {end_date_str}")
            report_lines.append(f"Data de Ontem (D-1): {end_date_str}")
            report_lines.append(f"Total de Lojas Analisadas: {total_stores_count}")
            report_lines.append("-"*80)
            
            report_lines.append("\n[MOVIMENTO 1] VERIFICAÇÃO DE ONTEM (D-1):")
            if not stores_missing_yesterday:
                report_lines.append("-> SUCESSO: Todas as lojas subiram as vendas de ontem com sucesso!")
            else:
                report_lines.append(f"-> ALERTA: {len(stores_missing_yesterday)} loja(s) NÃO subiram as vendas de ontem:")
                for store in sorted(stores_missing_yesterday):
                    meta = store_mapping.get(store, {"name": "N/A", "consultant": "N/A", "type": "N/A"})
                    report_lines.append(f"   - Loja {store} | {meta['name']} | Consultor: {meta['consultant']} | ({meta['type']})")
                    
            report_lines.append("\n[MOVIMENTO 2] VERIFICAÇÃO ACUMULADA DO MÊS (Dia 1 até Ontem):")
            stores_missing_month_list = sorted(stores_missing_month.items(), key=lambda x: x[1], reverse=True)
            if not stores_missing_month_list:
                report_lines.append("-> SUCESSO: Nenhuma loja possui pendências de subida neste mês.")
            else:
                report_lines.append(f"-> ALERTA: {len(stores_missing_month_list)} loja(s) possuem dias sem vendas acumulados no mês:")
                for store, days in stores_missing_month_list:
                    meta = store_mapping.get(store, {"name": "N/A", "consultant": "N/A", "type": "N/A"})
                    report_lines.append(f"   - Loja {store} | {meta['name']} | Consultor: {meta['consultant']} | ({meta['type']}) -> {days} dia(s) sem subir vendas")
                    
            report_lines.append("\n" + "="*80)
            
            report_content = "\n".join(report_lines)
            
            # Salvar Relatório TXT
            report_path = f"reports/relatorio_{date_suffix}.txt"
            with open(report_path, "w", encoding="utf-8") as f:
                f.write(report_content)
            logging.info(f"Relatório diário TXT salvo em {report_path}")
            
            # --- GERAR RELATÓRIO/PAINEL HTML ---
            generate_html_dashboard(stores_missing_yesterday, stores_missing_month, store_mapping, begin_date_str, end_date_str, total_stores_count)
            
            return True
            
        except Exception as e:
            logging.error(f"Erro na execução da automação: {e}", exc_info=True)
            try:
                await page.screenshot(path="screenshots/erro_execucao.png")
            except:
                pass
            return False
        finally:
            await browser.close()
            logging.info("Navegador fechado.")

if __name__ == "__main__":
    asyncio.run(run_automation())
